import io
import openai
import requests
import streamlit as st
import PyPDF2
from docx import Document
from openpyxl import load_workbook
import adal
from docx import Document


# SharePoint configuration
sharepoint_site_url = "https://your-sharepoint-site-url"
sharepoint_library_name = "YourLibraryName"
sharepoint_folder_name = "dataQandA"

# SharePoint OAuth configuration
client_id = "your_client_id"
client_secret = "your_client_secret"
resource = "https://<your-sharepoint-site-url>"
authority_url = "https://login.microsoftonline.com/<your-tenant-id>"
token_url = f"{authority_url}/oauth2/token"
# OpenAI API Key
openai.api_key = "your_openai_api_key"

# Function to obtain an access token using OAuth
def get_access_token():
    
    context = adal.AuthenticationContext(authority_url)
    token = context.acquire_token_with_client_credentials(resource, client_id, client_secret)
    if 'accessToken' in token:
        return token['accessToken']
    else:
        st.error("Error: Unable to obtain access token.")
        return None

# Function to retrieve a list of files from the specified SharePoint folder
def get_sharepoint_files(access_token):
    headers = {"Authorization": f"Bearer {access_token}"}
    api_url = f"{sharepoint_site_url}/_api/web/lists/getbytitle('{sharepoint_library_name}')/items?$filter=FileDirRef eq '{sharepoint_site_url}/{sharepoint_library_name}/{sharepoint_folder_name}'"
    response = requests.get(api_url, headers=headers)
    if response.status_code == 200:
        files = response.json()
        return files
    else:
        st.error("Error: Unable to retrieve SharePoint files.")
        return []

# Function to read text content from a PDF file
def read_pdf(file_url, access_token):
    pdf_text = ""
    try:
        with requests.get(file_url, headers={"Authorization": f"Bearer {access_token}"}, stream=True) as response:
            pdf_reader = PyPDF2.PdfFileReader(response.raw)
            for page_num in range(pdf_reader.getNumPages()):
                page = pdf_reader.getPage(page_num)
                pdf_text += page.extractText()
    except Exception as e:
        st.error(f"Error reading PDF: {str(e)}")
    return pdf_text

# Function to read text content from a Word document
def read_docx(file_url, access_token):
    docx_text = ""
    try:
        with requests.get(file_url, headers={"Authorization": f"Bearer {access_token}"}, stream=True) as response:
            doc = Document(io.BytesIO(response.content))
            for paragraph in doc.paragraphs:
                docx_text += paragraph.text
    except Exception as e:
        st.error(f"Error reading Word document: {str(e)}")
    return docx_text

# Function to read text content from an Excel file
def read_excel(file_url, access_token):
    excel_text = ""
    try:
        with requests.get(file_url, headers={"Authorization": f"Bearer {access_token}"}, stream=True) as response:
            workbook = load_workbook(io.BytesIO(response.content), read_only=True)
            for sheet in workbook:
                for row in sheet.iter_rows(values_only=True):
                    excel_text += " ".join(str(cell) for cell in row)
    except Exception as e:
        st.error(f"Error reading Excel file: {str(e)}")
    return excel_text

# Function to ask a question and generate an answer using GPT-3
def generate_answer(question, files):
    # Prepare the text for GPT-3
    context = "\n\n".join([f"{file['Title']}:\n{file['File_x0020_Type']}" for file in files])
    text = f"Context: {context}\n\nQuestion: {question}\nAnswer:"

    # Use GPT-3 to generate an answer
    response = openai.Completion.create(
    engine="text-davinci-002",
    prompt="Your prompt here",
    max_tokens=150
    )

    answer = response.choices[0].text
    return answer

# Main function
def main():
    # Obtain an access token
    access_token = get_access_token()

    # Retrieve files from the SharePoint folder
    sharepoint_files = get_sharepoint_files(access_token)

    # Ask a question
    user_question = input("Ask a question: ")

    # Generate an answer using GPT-3
    answer = generate_answer(user_question, sharepoint_files)

    print("Answer:", answer)

# Streamlit app layout
st.title("SharePoint Knowledge Base")

# Create a search box for user questions
user_question = st.text_input("Ask a question:")

# Create a button for searching
if st.button("Search"):
    # Ensure a question is provided
    if user_question:
        # Retrieve an access token for SharePoint
        access_token = get_access_token()
        if access_token:
            # Retrieve SharePoint files
            sharepoint_files = get_sharepoint_files(access_token)
            
            # Initialize variables for PDF, Word, and Excel content
            pdf_text = ""
            docx_text = ""
            excel_text = ""
            
            for file in sharepoint_files:
                file_url = file['ServerRelativeUrl']
                
                # Check the file extension to determine the file type
                if file_url.lower().endswith(".pdf"):
                    # Read content from a PDF file
                    pdf_text = read_pdf(file_url, access_token)
                elif file_url.lower().endswith(".docx"):
                    # Read content from a Word document
                    docx_text = read_docx(file_url, access_token)
                elif file_url.lower().endswith(".xlsx"):
                    # Read content from an Excel file
                    excel_text = read_excel(file_url, access_token)
            
            # Combine content from different file types for GPT-3 input
            combined_content = f"{pdf_text}\n{docx_text}\n{excel_text}"
            
            # Generate an answer using GPT-3
            answer = generate_answer(user_question, combined_content)
            
            # Display the answer
            st.subheader("Answer:")
            st.write(answer)

            # Add a feedback mechanism
            feedback = st.selectbox("Was this answer helpful?", ["Yes", "No"])
            if feedback == "No":
                issue_description = st.text_area("Please describe the issue:")
                if st.button("Submit Feedback"):
                    # Implement feedback submission logic here
                    st.success("Thank you for your feedback! We will review it.")
        else:
            st.warning("Please check SharePoint OAuth configuration.")
    else:
        st.warning("Please enter a question to search.")

if __name__ == "__main__":
    main()
